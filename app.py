"""
全能扫描系统 - Flask主应用
路由、API端点、启动入口
"""

import os
import sys
import socket
import logging
from io import BytesIO

from flask import Flask, request, jsonify, render_template, send_file

# 抑制 Flask/Werkzeug 开发服务器警告
logging.getLogger('werkzeug').setLevel(logging.ERROR)
cli = sys.modules.get('flask.cli')
if cli:
    cli.show_server_banner = lambda *args, **kwargs: None
import qrcode
from openpyxl import load_workbook, Workbook

from database import (
    init_db, get_all_datasets, get_dataset, get_current_dataset,
    set_current_dataset, delete_dataset, create_dataset, insert_boxes,
    find_box, mark_scanned, add_scan_record, get_progress,
    get_recent_scan_records, get_all_boxes
)

app = Flask(__name__)
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16MB
app.config['SERVER_PORT'] = 5000


# ========== Utilities ==========

def get_local_ip():
    """获取本机局域网IP地址"""
    try:
        s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
        s.connect(("8.8.8.8", 80))
        ip = s.getsockname()[0]
        s.close()
        return ip
    except Exception:
        return "127.0.0.1"


def parse_excel(file_stream):
    """
    解析上传的Excel文件
    返回: list of tuples [(box_number, store_address, zone), ...]
    """
    wb = load_workbook(file_stream, read_only=True)
    ws = wb.active

    boxes = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:  # 跳过表头行
            continue
        if not row or len(row) < 3:
            continue

        box_number = str(row[0]).strip() if row[0] else ''
        store_address = str(row[1]).strip() if row[1] else ''
        zone = str(row[2]).strip() if row[2] else ''

        if box_number:
            boxes.append((box_number, store_address, zone))

    wb.close()
    return boxes


def generate_export(dataset_name, boxes):
    """
    生成导出Excel
    返回: BytesIO 对象
    """
    wb = Workbook()
    ws = wb.active
    ws.title = '扫描结果'

    ws.append(['箱号', '门店地址', '分区', '扫描状态', '扫描时间'])

    for box in boxes:
        ws.append([
            box['box_number'],
            box['store_address'],
            box['zone'],
            '已扫描' if box['is_scanned'] else '未扫描',
            box['scanned_at'] or ''
        ])

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


# ========== Page Routes ==========

@app.route('/')
def index():
    """管理页面"""
    return render_template('index.html')


# ========== API Routes ==========

@app.route('/api/ping')
def ping():
    """连接测试接口"""
    return jsonify({'status': 'ok', 'message': 'connected'})


@app.route('/api/qrcode')
def get_qrcode():
    """生成连接二维码"""
    server_url = f"http://{get_local_ip()}:{app.config['SERVER_PORT']}"
    qr = qrcode.QRCode(version=1, box_size=10, border=5)
    qr.add_data(server_url)
    qr.make(fit=True)
    img = qr.make_image(fill_color="black", back_color="white")

    img_io = BytesIO()
    img.save(img_io, 'PNG')
    img_io.seek(0)

    return send_file(img_io, mimetype='image/png')


@app.route('/api/upload', methods=['POST'])
def upload_file():
    """上传Excel数据集"""
    # 验证文件
    if 'file' not in request.files:
        return jsonify({'success': False, 'message': '请选择文件'}), 400

    file = request.files['file']
    if file.filename == '':
        return jsonify({'success': False, 'message': '请选择文件'}), 400

    if not file.filename.endswith('.xlsx'):
        return jsonify({'success': False, 'message': '仅支持.xlsx格式文件'}), 400

    # 验证名称
    name = request.form.get('name', '').strip()
    if not name:
        return jsonify({'success': False, 'message': '请输入数据集名称'}), 400

    # 解析Excel
    try:
        boxes = parse_excel(file)
    except Exception:
        return jsonify({'success': False, 'message': '文件格式不正确，无法解析'}), 400

    if not boxes:
        return jsonify({'success': False, 'message': '文件中没有数据'}), 400

    # 写入数据库
    dataset_id = create_dataset(name, file.filename, len(boxes))
    insert_boxes(dataset_id, boxes)

    return jsonify({
        'success': True,
        'message': '导入成功',
        'dataset': {
            'id': dataset_id,
            'name': name,
            'total_count': len(boxes)
        }
    })


@app.route('/api/datasets')
def list_datasets():
    """获取数据集列表"""
    datasets = get_all_datasets()
    return jsonify({'datasets': datasets})


@app.route('/api/datasets/<int:dataset_id>', methods=['DELETE'])
def remove_dataset(dataset_id):
    """删除数据集"""
    if delete_dataset(dataset_id):
        return jsonify({'success': True, 'message': '删除成功'})
    return jsonify({'success': False, 'message': '数据集不存在'}), 404


@app.route('/api/set-current-dataset', methods=['POST'])
def set_current():
    """设置当前数据集"""
    data = request.get_json()
    if not data or 'dataset_id' not in data:
        return jsonify({'success': False, 'message': '缺少dataset_id参数'}), 400

    if set_current_dataset(data['dataset_id']):
        return jsonify({'success': True, 'message': '设置成功'})
    return jsonify({'success': False, 'message': '数据集不存在'}), 404


@app.route('/api/current-dataset')
def current_dataset():
    """获取当前数据集"""
    ds = get_current_dataset()
    if ds:
        return jsonify(ds)
    return jsonify({'success': False, 'message': '未设置当前数据集'}), 404


@app.route('/api/scan', methods=['POST'])
def scan_box():
    """扫描箱号查询"""
    data = request.get_json()
    if not data:
        return jsonify({'success': False, 'message': '无效请求'}), 400

    dataset_id = data.get('dataset_id')
    box_number = data.get('box_number', '').strip()
    device_id = data.get('device_id')

    if not dataset_id or not box_number:
        return jsonify({'success': False, 'message': '缺少必要参数'}), 400

    box = find_box(dataset_id, box_number)

    if not box:
        # 未找到
        add_scan_record(dataset_id, box_number, '', '', 'not_found', device_id)
        return jsonify({
            'status': 'not_found',
            'box_number': box_number
        })

    first_scan = False
    if box['is_scanned'] == 0:
        # 首次扫描
        mark_scanned(box['id'], dataset_id, device_id)
        first_scan = True
        status = 'normal'
    else:
        # 重复扫描
        status = 'duplicate'

    # 添加扫描记录
    add_scan_record(dataset_id, box_number, box['zone'], box['store_address'], status, device_id)

    # 获取最新进度
    progress = get_progress(dataset_id)

    return jsonify({
        'status': 'found',
        'box_number': box_number,
        'zone': box['zone'],
        'store_address': box['store_address'],
        'first_scan': first_scan,
        'progress': progress
    })


@app.route('/api/scan/recent')
def recent_records():
    """获取最近扫描记录"""
    limit = request.args.get('limit', 20, type=int)
    records = get_recent_scan_records(limit)
    return jsonify({'records': records})


@app.route('/api/scan/progress/<int:dataset_id>')
def scan_progress(dataset_id):
    """获取扫描进度"""
    return jsonify(get_progress(dataset_id))


@app.route('/api/export/<int:dataset_id>')
def export_dataset(dataset_id):
    """导出Excel"""
    ds = get_dataset(dataset_id)
    if not ds:
        return jsonify({'success': False, 'message': '数据集不存在'}), 404

    boxes = get_all_boxes(dataset_id)
    output = generate_export(ds['name'], boxes)

    filename = f"{ds['name']}_导出.xlsx"

    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )


# ========== Error Handlers ==========

@app.errorhandler(413)
def too_large(e):
    return jsonify({'success': False, 'message': '文件大小超过限制（最大16MB）'}), 413


@app.errorhandler(Exception)
def handle_error(error):
    return jsonify({'success': False, 'message': '服务器内部错误'}), 500


# ========== Startup ==========

if __name__ == '__main__':
    os.makedirs('data', exist_ok=True)
    init_db()

    port = app.config['SERVER_PORT']
    local_ip = get_local_ip()

    print('=' * 48)
    print('        全能扫描系统 已启动')
    print('=' * 48)
    print(f'  本地访问：http://localhost:{port}')
    print(f'  局域网访问：http://{local_ip}:{port}')
    print()
    print('  手持终端请扫描二维码或输入以上地址连接')
    print('  按 Ctrl+C 停止服务')
    print('=' * 48)

    app.run(host='0.0.0.0', port=port, debug=False)
